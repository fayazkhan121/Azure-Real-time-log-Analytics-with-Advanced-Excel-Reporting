import os
from azure.identity import DefaultAzureCredential
from azure.mgmt.monitor import MonitorManagementClient
from azure.mgmt.subscription import SubscriptionClient
from azure.eventhub import EventHubConsumerClient
from azure.storage.blob import BlobServiceClient
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference

# Configuration
TENANT_ID = os.environ['AZURE_TENANT_ID']
SUBSCRIPTION_ID = os.environ['AZURE_SUBSCRIPTION_ID']
RESOURCE_GROUP = 'your-resource-group-name'
EVENT_HUB_NAMESPACE = 'your-eventhub-namespace'
EVENT_HUB_NAME = 'insights-logs-activitylogs'
STORAGE_ACCOUNT_NAME = 'your-storage-account-name'
CONTAINER_NAME = 'log-analysis-reports'

# Clients
credential = DefaultAzureCredential()
monitor_client = MonitorManagementClient(credential, SUBSCRIPTION_ID) 
subscription_client = SubscriptionClient(credential)
consumer_client = EventHubConsumerClient.from_connection_string(
    os.environ['EVENT_HUB_CONNECTION_STRING'],
    eventhub_name=EVENT_HUB_NAME,  
    consumer_group='$Default'
)
storage_account_url = f"https://{STORAGE_ACCOUNT_NAME}.blob.core.windows.net"
blob_service_client = BlobServiceClient(storage_account_url, credential)

# Main processing
def main():
    try:
        # Get all subscriptions
        subscriptions = subscription_client.subscriptions.list()

        # Ingest real-time logs for each subscription
        log_data = {}
        for subscription in subscriptions:
            log_data[subscription.subscription_id] = ingest_subscription_logs(subscription.subscription_id)

        # Analyze and process log data
        analysis_results = process_log_data(log_data)

        # Generate advanced Excel report
        generate_excel_report(log_data, analysis_results)

    except Exception as e:
        print(f"Log analysis failed: {e}")

# Ingest logs from Azure Event Hub
def ingest_subscription_logs(subscription_id):
    print(f"Ingesting real-time logs for subscription: {subscription_id}")
    df = pd.DataFrame(columns=['TimeGenerated', 'ResourceId', 'OperationName', 'Level', 'Message'])

    def on_event(partition_context, event):
        event_dict = {
            'TimeGenerated': event.message.annotations['microsoft.azure.monitor.metricValueTimeUTC'],
            'ResourceId': event.message.annotations['resourceId'],
            'OperationName': event.message.annotations['operationName'],
            'Level': event.message.annotations['level'],
            'Message': event.body_as_str(encoding='UTF-8')
        }
        df.loc[len(df)] = event_dict  

    with consumer_client:
        consumer_client.receive(on_event=on_event, starting_position="-1")
    
    return df

# Analyze and process log data
def process_log_data(log_data):
    analysis_results = {}

    for sub_id, sub_df in log_data.items():
        print(f"Analyzing logs for subscription: {sub_id}")

        # Enrich with resource metadata
        resource_ids = sub_df['ResourceId'].unique()
        sub_resources = monitor_client.resources.list(sub_id, expand="tags")
        resource_df = pd.DataFrame(data=[(r.id, r.resource_group, r.location, r.tags) for r in sub_resources],
                                columns=['ResourceId', 'ResourceGroup', 'Location', 'Tags'])
        sub_df = sub_df.merge(resource_df, on='ResourceId', how='left')

        # Get operation category
        sub_df['Category'] = sub_df['OperationName'].apply(lambda x: x.split('/')[0])

        # Summarize metrics
        op_count = sub_df.groupby('OperationName').size().reset_index(name='Count')
        sev_count = sub_df.groupby('Level').size().reset_index(name='Count')
        res_count = sub_df.groupby('ResourceId').size().reset_index(name='Count')

        analysis_results[sub_id] = {
            'subscription_logs': sub_df, 
            'operation_summary': op_count,
            'severity_summary': sev_count,  
            'resource_summary': res_count
        }

    return analysis_results

# Generate Excel report
def generate_excel_report(log_data, analysis_results):
    print("Generating Excel report")
    wb = Workbook()
    del wb['Sheet']

    # Formats
    header_font = Font(size=12, bold=True)
    table_header_fill = PatternFill("solid", fgColor="5FA8D3") 
    table_header_font = Font(color="FFFFFF", bold=True)
    table_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Summary sheet
    print("Creating summary sheet")
    ws_summary = wb.create_sheet('Log Analysis Summary')
    ws_summary.column_dimensions['A'].width = 25
    ws_summary['A1'] = "Azure Log Analysis Summary"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    start_row = 3
    ws_summary.cell(row=start_row, column=1, value="Subscription ID") 
    ws_summary.cell(row=start_row, column=2, value="Total Logs")
    ws_summary.cell(row=start_row, column=3, value="Error Logs")
    ws_summary.cell(row=start_row, column=4, value="Resource Types")
    for cell in ws_summary[start_row]:
        cell.font = header_font
    start_row += 1

    for sub_id, sub_data in analysis_results.items():  
        ws_summary.cell(row=start_row, column=1, value=sub_id)
        ws_summary.cell(row=start_row, column=2, value=len(sub_data['subscription_logs']))
        ws_summary.cell(row=start_row, column=3, value=len(sub_data['subscription_logs'][sub_data['subscription_logs']['Level']=='Error']))
        ws_summary.cell(row=start_row, column=4, value=sub_data['subscription_logs']['ResourceId'].nunique())
        start_row += 1

    # Subscription sheets  
    for sub_id, sub_data in analysis_results.items():
        print(f"Creating sheet for subscription: {sub_id}")

        # Raw data
        ws_raw = wb.create_sheet(f"{sub_id}_RawData")  
        for r in dataframe_to_rows(sub_data['subscription_logs'], index=False, header=True):
            ws_raw.append(r)
        for cell in ws_raw[1]:
            cell.fill = table_header_fill 
            cell.font = table_header_font
            cell.alignment = table_header_alignment

        # Analysis
        ws_analysis = wb.create_sheet(f"{sub_id}_Analysis")
        
        # Top operations
        op_start_row = 1
        ws_analysis.cell(row=op_start_row, column=1, value="Top Operations").font = header_font
        op_start_row += 1
        for r in dataframe_to_rows(sub_data['operation_summary'].head(5), index=False, header=True):
            ws_analysis.append(r)
        for cell in ws_analysis[op_start_row]:
            cell.fill = table_header_fill
            cell.font = table_header_font  
            cell.alignment = table_header_alignment
        
        # Log severity  
        sev_start_row = op_start_row + sub_data['operation_summary'].head(5).shape[0] + 2
        ws_analysis.cell(row=sev_start_row, column=1, value="Log Severity").font = header_font
        sev_start_row += 1  
        for r in dataframe_to_rows(sub_data['severity_summary'], index=False, header=True):
            ws_analysis.append(r)
        for cell in ws_analysis[sev_start_row]:
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = table_header_alignment

        # Charts
        # Severity pie chart
        severity_chart = PieChart()
        severity_chart.title = f"Log Severity Breakdown - {sub_id}"
        labels = Reference(ws_analysis, min_col=1, min_row=sev_start_row+1, max_row=ws_analysis.max_row)
        data = Reference(ws_analysis, min_col=2, min_row=sev_start_row, max_row=ws_analysis.max_row)
        severity_chart.add_data(data, titles_from_data=True)
        severity_chart.set_categories(labels)
        severity_chart.width = 12
        severity_chart.height = 8
        ws_analysis.add_chart(severity_chart, "E3") 

        # Operation bar chart
        op_chart = BarChart()  
        op_chart.title = f"Top Operations - {sub_id}"
        op_chart.y_axis.title = 'Operation'
        op_chart.x_axis.title = 'Count'
        op_data = Reference(ws_analysis, min_col=2, min_row=op_start_row, max_row=op_start_row+4)
        op_cats = Reference(ws_analysis, min_col=1, min_row=op_start_row+1, max_row=op_start_row+5)
        op_chart.add_data(op_data, titles_from_data=True)
        op_chart.set_categories(op_cats)
        op_chart.width = 20
        op_chart.height = 8
        ws_analysis.add_chart(op_chart, "E18")

    # Save the report
    filename = f"Azure_Log_Analysis_{pd.Timestamp('now').strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    blob_client = blob_service_client.get_blob_client(container=CONTAINER_NAME, blob=filename)
    with blob_client.stage_block() as block:
        stream = io.BytesIO()
        wb.save(stream)  
        block.upload_blob(stream.getvalue())

    print(f"Report '{filename}' uploaded to Azure Storage.")

if __name__ == "__main__":
    main()
